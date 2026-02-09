from flask import Blueprint, jsonify, request
from app.models.user import User
from app.models.class_model import Class
from app.models.face_update_request import FaceUpdateRequest
from flask_jwt_extended import jwt_required, get_jwt_identity
from datetime import datetime

bp = Blueprint('admin', __name__, url_prefix='/api/admin')

@bp.route('/stats', methods=['GET'])
@jwt_required()
def get_stats():
    current_user_id = get_jwt_identity()
    current_user = User.query.get(int(current_user_id))
    
    if not current_user or current_user.role != 'admin':
        return jsonify({'message': 'Access forbidden: Admins only'}), 403

    # In a real app, check for admin role here
    total_users = User.query.count()
    total_classes = Class.query.count()
    active_students = User.query.filter_by(role='student', is_active=True).count()
    
    return jsonify({
        'total_users': total_users,
        'total_classes': total_classes,
        'active_students': active_students
    }), 200

# --- User CRUD ---

@bp.route('/users', methods=['GET'])
@jwt_required()
def get_users():
    current_user_id = get_jwt_identity()
    current_user = User.query.get(int(current_user_id))
    if not current_user or current_user.role != 'admin':
         return jsonify({'message': 'Access forbidden'}), 403

    role_filter = request.args.get('role')
    query = User.query
    if role_filter:
        query = query.filter_by(role=role_filter)
    
    users = query.all()
    return jsonify([u.to_dict() for u in users]), 200

@bp.route('/users', methods=['POST'])
@jwt_required()
def create_user():
    current_user_id = get_jwt_identity()
    current_user = User.query.get(int(current_user_id))
    if not current_user or current_user.role != 'admin':
         return jsonify({'message': 'Access forbidden'}), 403

    data = request.get_json()
    
    if User.query.filter_by(username=data.get('username')).first():
        return jsonify({'message': 'Username already exists'}), 400
    if User.query.filter_by(email=data.get('email')).first():
        return jsonify({'message': 'Email already exists'}), 400

    new_user = User(
        username=data.get('username'),
        email=data.get('email'),
        role=data.get('role', 'student'),
        full_name=data.get('full_name')
    )
    new_user.set_password(data.get('password'))
    
    from app.extensions import db
    db.session.add(new_user)
    db.session.commit()
    
    return jsonify({'message': 'User created successfully', 'user': new_user.to_dict()}), 201

@bp.route('/users/<int:user_id>', methods=['PUT'])
@jwt_required()
def update_user(user_id):
    current_user_id = get_jwt_identity()
    current_user = User.query.get(int(current_user_id))
    if not current_user or current_user.role != 'admin':
         return jsonify({'message': 'Access forbidden'}), 403

    user = User.query.get_or_404(user_id)
    data = request.get_json()
    
    if 'full_name' in data:
        user.full_name = data['full_name']
    if 'email' in data:
        user.email = data['email']
    if 'role' in data:
        user.role = data['role']
    if 'password' in data and data['password']:
        user.set_password(data['password'])
        
    from app.extensions import db
    db.session.commit()
    return jsonify({'message': 'User updated successfully', 'user': user.to_dict()}), 200

@bp.route('/users/<int:user_id>', methods=['DELETE'])
@jwt_required()
def delete_user(user_id):
    current_user_id = get_jwt_identity()
    current_user = User.query.get(int(current_user_id))
    if not current_user or current_user.role != 'admin':
         return jsonify({'message': 'Access forbidden'}), 403

    user = User.query.get_or_404(user_id)
    
    from app.extensions import db
    db.session.delete(user)
    db.session.commit()
    return jsonify({'message': 'User deleted successfully'}), 200


# --- Face Update Requests ---

@bp.route('/face-update-requests', methods=['GET'])
@jwt_required()
def get_face_update_requests():
    current_user_id = get_jwt_identity()
    current_user = User.query.get(int(current_user_id))
    if not current_user or current_user.role != 'admin':
        return jsonify({'message': 'Access forbidden'}), 403

    status_filter = request.args.get('status')
    query = FaceUpdateRequest.query
    if status_filter:
        query = query.filter_by(status=status_filter)
    
    requests = query.order_by(FaceUpdateRequest.created_at.desc()).all()
    return jsonify([r.to_dict() for r in requests]), 200

@bp.route('/face-update-requests/<int:request_id>/approve', methods=['POST'])
@jwt_required()
def approve_face_update_request(request_id):
    current_user_id = get_jwt_identity()
    current_user = User.query.get(int(current_user_id))
    if not current_user or current_user.role != 'admin':
        return jsonify({'message': 'Access forbidden'}), 403

    update_request = FaceUpdateRequest.query.get_or_404(request_id)
    
    if update_request.status != 'pending':
        return jsonify({'message': 'Request already processed'}), 400
    
    # Update request status
    update_request.status = 'approved'
    update_request.reviewed_by = current_user.id
    update_request.reviewed_at = datetime.utcnow()
    
    # Grant permission to user
    user = User.query.get(update_request.user_id)
    user.face_update_allowed = True
    
    from app.extensions import db
    db.session.commit()
    
    return jsonify({'message': 'Request approved. User can now update their face.'}), 200

@bp.route('/face-update-requests/<int:request_id>/deny', methods=['POST'])
@jwt_required()
def deny_face_update_request(request_id):
    current_user_id = get_jwt_identity()
    current_user = User.query.get(int(current_user_id))
    if not current_user or current_user.role != 'admin':
        return jsonify({'message': 'Access forbidden'}), 403

    update_request = FaceUpdateRequest.query.get_or_404(request_id)
    
    if update_request.status != 'pending':
        return jsonify({'message': 'Request already processed'}), 400
    
    # Update request status
    update_request.status = 'denied'
    update_request.reviewed_by = current_user.id
    update_request.reviewed_at = datetime.utcnow()
    
    from app.extensions import db
    db.session.commit()
    
    return jsonify({'message': 'Request denied.'}), 200



# --- Export Attendance to Excel ---
@bp.route('/export-attendance', methods=['GET'])
@jwt_required()
def export_attendance():
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from app.models.attendance import Attendance
    from app.models.class_model import AttendanceSession
    import io
    
    current_user_id = get_jwt_identity()
    current_user = User.query.get(int(current_user_id))
    if not current_user or current_user.role != 'admin':
        return jsonify({'message': 'Access forbidden'}), 403
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance Records'
    
    # Header style
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    # Headers
    headers = ['Student ID', 'Student Name', 'Email', 'Class', 'Subject', 'Date', 'Time', 'Status', 'Confidence Percent']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Fetch all attendance records
    attendances = Attendance.query.join(User).join(AttendanceSession).all()
    
    # Add data rows
    for att in attendances:
        confidence_pct = round(att.confidence_score * 100, 2) if att.confidence_score else 'N/A'
        ws.append([
            att.student.id,
            att.student.full_name or att.student.username,
            att.student.email,
            att.session.class_obj.name if att.session.class_obj else 'N/A',
            att.session.class_obj.code if att.session.class_obj else 'N/A',
            att.timestamp.strftime('%Y-%m-%d'),
            att.timestamp.strftime('%H:%M:%S'),
            att.status,
            confidence_pct
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='attendance_records.xlsx'
    )

